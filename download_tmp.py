from header_relevant import send_headers
from openpyxl.styles import Font, Alignment
from read_file import read_pdf, read_docx, read_rtf, read_xlsx, read_csv, read_pptx, read_txt
from win32com.client import Dispatch
import openpyxl
import requests
import csv
import io
import os
import sys
import time


def info_write(file_type, target, search_rusult, numbur, title, url):
    rusult_count = '符合特徵：\n姓名 {} 個\n身分證 {} 個\n電話 {} 個\n信箱 {} 個\n地址 {} 個\n生日 {} 個\n'.format(len(search_rusult["username"]), len(
        search_rusult["id"]), len(search_rusult["ph_no"]), len(search_rusult["em"]), len(search_rusult["address"]), len(search_rusult["bir"]))
    remark = '姓名：{}\n身分證：{}\n電話：{}\n信箱：{}\n地址：{}\n生日：{}'.format(
        search_rusult["username"], search_rusult["id"], search_rusult["ph_no"], search_rusult["em"], search_rusult["address"], search_rusult["bir"])
    sum_count = len(search_rusult["username"]) + len(search_rusult["id"]) + len(search_rusult["ph_no"]) + len(
        search_rusult["em"]) + len(search_rusult["address"]) + len(search_rusult["bir"])
    if sum_count != 0:
        if file_type == 'csv':
            with open(target + '_info.csv', 'a+', encoding='utf-8-sig', newline='') as fcsv:
                writer = csv.writer(fcsv)
                writer.writerow(
                    [numbur+1, title, rusult_count, url, remark])
            fcsv.close()
        elif file_type == 'xlsx':
            wb = openpyxl.load_workbook(target + '_info.xlsx')
            sheet = wb[target]
            data = [sheet.max_row-2, title, rusult_count, url, remark]
            row = sheet.max_row+1
            for i in range(5):
                sheet.cell(row, i+1).value = data[i]
                sheet.cell(row, i+1).font = Font(name=u'微軟正黑體', size=12)
                sheet.cell(row, i+1).alignment = Alignment(wrapText=True,
                                                           vertical="center")  # 自動換行, 置中對齊
            wb.save(target + '_info.xlsx')
        print(f"\rFind info {sum_count} count.")
        print("\r")
    else:
        print("\rNot find info.")
        print("\r")


def detcet_type(file_type, target, count, titles, links):
    for i in range(count):
        print(f"\r[{i+1}] Title :", titles[i])
        print("\rUrl :", links[i])
        url_to_string = str(links[i])
        title_to_string = str(titles[i])
        response = requests.get(url_to_string, headers=send_headers)
        bytes_io = io.BytesIO(response.content)
        ext = url_to_string.split('.')[-1]
        # 判斷下載檔案副檔名
        if ext == "pdf":
            store_path = './tmp/tmp.' + ext
            # 下載檔案
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            # print('pdftmp.pdf, download success!')
            search_rusult = read_pdf(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "docx":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_docx(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "doc":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            # 將doc轉乘docx
            word = Dispatch("Word.Application")
            wordfile = word.Documents.Open(sys.path[0] + store_path)
            # 另存新檔,附檔名多添加x,儲存成docx格式代碼
            wordfile.SaveAs(
                sys.path[0] + store_path + 'x', FileFormat=12)
            wordfile.Close()
            word.Quit()
            os.remove(store_path)
            search_rusult = read_docx(store_path+'x')
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "rtf" or ext == "odt":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_rtf(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "xlsx" or ext == "xls":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_xlsx(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "csv":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_csv(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "pptx":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_pptx(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "ppt":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            # 將ppt轉乘pptx
            ppt = Dispatch("Powerpoint.Application")
            pptfile = ppt.Presentations.Open(
                sys.path[0] + store_path, WithWindow=False)
            # 另存新檔,附檔名多添加x,儲存成pptx格式代碼
            pptfile.SaveAs(
                sys.path[0] + store_path + 'x', FileFormat=24)
            pptfile.Close()
            ppt.Quit()
            os.remove(store_path)
            search_rusult = read_pptx(store_path+'x')
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        elif ext == "txt":
            store_path = './tmp/tmp.' + ext
            with open(store_path, mode="wb") as f:
                f.write(bytes_io.getvalue())
            f.close()
            search_rusult = read_txt(store_path)
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        else:
            print(f"\rFilename Extension : {ext} Not support!")
            search_rusult = {"username": [], "id": [], "ph_no": [],
                             "em": [], "address": [], "bir": []}
            info_write(file_type, target, search_rusult, i,
                       title_to_string, url_to_string)
        time.sleep(1)
