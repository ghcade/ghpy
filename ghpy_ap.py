from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from check_input import check_input
from openpyxl.styles import Font, Alignment
from header_relevant import header_list
from download_tmp import detcet_type
import requests
import openpyxl
import threading
import csv
import sys
import time
import random

sys.tracebacklimit = 0


def deley_process(page):
    n = random.randint(10, 25)
    for i in range(n):
        time.sleep(1)
        sys.stdout.write(
            '\r' + '            ' + ' Page {} , '.format(page) + ' Google search... {:.2f}'.format(i/n*100)+'%')
        sys.stdout.flush()
    print('\rPage {} , Google search finished!                                   '.format(page))
    print('\r')
    time.sleep(2)


def reoprt_process():
    chars = "/-\|"
    #chars = "⢿⣻⣽⣾⣷⣯⣟⡿"
    for char in chars:
        sys.stdout.write('\r'+'Comparing...  ' + char)
        time.sleep(.1)
        sys.stdout.flush()


def animated_loading():
    #self.steps = ["⢿", "⣻", "⣽", "⣾", "⣷", "⣯", "⣟", "⡿"]
    #chars = "⢿⣻⣽⣾⣷⣯⣟⡿"
    #chars = "/—\|"
    chars = "/-\|"
    #chars = [".", "..", "..."]
    for char in chars:
        sys.stdout.write('\r'+'Executing ' + char)
        time.sleep(.1)
        sys.stdout.flush()


def google_search(tg, ft):
    #site = "site:www.bbl.com.tw"
    #keyword = """+("身份證字號")"""
    #ext = "+ext:doc"
    keyword = """+("證字號"+|+"姓名"+|+"生日"+|+"出生"+|+"電話"+|+"手機"+|+"護照"+|+"聯絡")"""
    ext = "+ext:doc+|+ext:docx+|+ext:xls+|+ext:xlsx+|+ext:ppt+|+ext:pptx+|+ext:pdf+|+ext:csv+|+ext:odt+|+ext:rtf"
    query = "site:" + tg + keyword + ext  # 查詢條件
    links = []  # 存url的list
    titles = []  # 存title的list
    search_list = ['yuRUbf', 'egMi0 kCrYT']  # 爬蟲Url可能擷取的class
    check_list = ['http://', 'https://']  # 因class擷取的資訊可能不同，已http(s)來確認擷取內容
    count = 0  # 統計搜尋總數
    page = 1  # 搜尋頁數
    end = 0  # 是否搜尋完畢flag
    now_time = time.strftime(r"%Y-%m-%d %H:%M", time.localtime())  # 時間處理
    print('\rStarting GHPY at {} CST'.format(now_time))  # 印出時間

    s = open('proxy_list.txt', 'r', encoding='UTF-8')
    proxy_sum = len(s.readlines())  # free proxy 總數
    s.close()
    proxy_no = 0  # 紀錄使用到第幾個 proxy
    prxoy_ip = ''
    while end == 0:
        try:
            user_agent = random.choice(header_list)  # 隨機user_agent
            chrome_options = Options()
            chrome_options.binary_location = "C:\Program Files\Google\Chrome\Application\chrome.exe"  # chrome 位置
            # chrome_options.add_argument("--headless")  # 隱藏視窗
            chrome_options.add_argument('--incognito')  # 使用無痕視窗
            chrome_options.add_argument(f'user-agent={user_agent}')
            if prxoy_ip != '':
                chrome_options.add_argument(f'--proxy-server={prxoy_ip}')
            chrome_options.add_experimental_option(
                'excludeSwitches', ['enable-logging'])  # 隱藏chrome在command的log

            driver = webdriver.Chrome(options=chrome_options)
            url = "https://www.google.com/search?q=" + \
                query + "&start=" + str((page - 1) * 10) + "&filter=0"
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            check_rusult_flag = 0
            if soup.find_all(id="recaptcha") and proxy_no < proxy_sum:
                print("\rPage {} , Verification required!".format(page))
                try:
                    #proxies = {'https': []}
                    pl = open('proxy_list.txt', 'r',
                              encoding='UTF-8')  # blacklist
                    line = pl.readlines()[proxy_no].replace('\n', '')
                    prxoy_ip = line
                    #proxies['https'] = line
                    pl.close()
                    #prxoy_ip = requests.get('https://api.ipify.org', proxies=proxies).text
                    print("\rChange IP : {} ".format(prxoy_ip))
                    proxy_no += 1
                    time.sleep(1)
                except Exception as e:
                    print(e)
                    proxy_no += 1
                    pass
            elif soup.find_all('div', class_='error-code') and proxy_no < proxy_sum:
                print("\rPage {} , Google TimeOut!".format(page))
                try:
                    #proxies = {'https': []}
                    pl = open('proxy_list.txt', 'r',
                              encoding='UTF-8')  # blacklist
                    line = pl.readlines()[proxy_no].replace('\n', '')
                    prxoy_ip = line
                    #proxies['https'] = line
                    pl.close()
                    #prxoy_ip = requests.get('https://api.ipify.org', proxies=proxies).text
                    print("\rChange IP : {} ".format(prxoy_ip))
                    proxy_no += 1
                    time.sleep(1)
                except Exception as e:
                    print(e)
                    proxy_no += 1
                    pass
            elif proxy_no > proxy_sum:
                print("\rPage {} , Verification required!".format(page))
                print("\rProxy is over!")
                end = 1
            else:
                for c in search_list:
                    search = soup.find_all('div', c)
                    if len(search) != 0:
                        for lt in search:
                            title = lt.a.h3.text
                            link = lt.a.get('href')
                            for check_url in check_list:
                                if link.find(check_url) != -1:
                                    link = check_url+tg+link.split(tg)[1]
                                    link = link.split('&ved=')[0]
                            try:
                                titles.append(title)
                                links.append(link)
                            except Exception as e:
                                print("\rError log : " + e)
                                pass
                            count += 1  # 總數
                            check_rusult_flag += 1  # 搜尋結果
                        deley_process(page)
                        page += 1
                        # time.sleep(random.randint(10, 25))  # 時間延遲
                        break
                if check_rusult_flag == 0:
                    end = 1
        except Exception as e:
            print(e)
            pass

    if ft == 'csv':
        with open(tg + '_info.csv', 'w', encoding='utf-8-sig', newline='') as fcsv:
            writer = csv.writer(fcsv)
            report_time = 'Starting GHPY at {} CST'.format(now_time)
            writer.writerow(['網頁個資檢測 {} '.format(tg)])
            writer.writerow([report_time])  # 寫入時間
            writer.writerow(['No', 'Title', 'Rusult', 'URL', 'Remark'])
        fcsv.close()
        # reoprt_process()
        detcet_type(ft, tg, count, titles, links)
    elif ft == 'xlsx':
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        report_time = 'Starting GHPY at {} CST'.format(now_time)
        sheet['A1'].value = '網頁個資檢測 {} '.format(tg)
        sheet['A1'].font = Font(name=u'微軟正黑體', size=14, bold=True)
        sheet['A2'].value = report_time
        sheet['A2'].font = Font(name=u'微軟正黑體', size=14, bold=True)
        data = ['No', 'Title', 'Rusult', 'URL', 'Remark']
        for i in range(5):
            sheet.cell(3, i+1).value = data[i]
            sheet.cell(3, i+1).alignment = Alignment(horizontal='center')
            sheet.cell(3, i+1).font = Font(name=u'微軟正黑體', size=12)
        # 合併儲存格
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        # 置中
        sheet.cell(1, 1).alignment = Alignment(horizontal='center')
        sheet.cell(2, 1).alignment = Alignment(horizontal='center')
        # 寬度調整
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 45
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 45
        sheet.column_dimensions['E'].width = 30
        # sheet名稱
        sheet.title = tg
        wb.save(tg + '_info.xlsx')
        # reoprt_process()
        detcet_type(ft, tg, count, titles, links)


if __name__ == "__main__":
    try:
        #target_url, file_type = check_input(sys.argv[1:5])
        target_url = 'www.yuantabank.com.tw'
        file_type = 'xlsx'
        ip = requests.get('https://api.ipify.org').text
        print("\rUse IP : {}".format(ip))
        the_process = threading.Thread(
            target=google_search, args=(target_url, file_type))
        the_process.start()
        time.sleep(1)
        while the_process.is_alive():
            animated_loading()
        sys.stdout.write(f'\rDone!                                     ')
    except Exception as e:
        print(e)
        #print('Use : ghpy.py -u <target url> [-t <file type>]')
